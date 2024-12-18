import os
import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime


def read_file(file_name):
    """Read items from a text file into a dictionary {id: name}."""
    if not os.path.exists(file_name):
        return {}
    with open(file_name, "r") as file:
        return {
            line.split()[0]: " ".join(line.split()[1:])
            for line in file if line.strip()
        }

def write_file(file_name, items):
    """Write items to a text file from a dictionary {id: name}."""
    with open(file_name, "w") as file:
        for user_id, username in items.items():
            file.write(f"{user_id} {username}\n")

def add_to_file(file_name, user_id, username):
    """Add an entry to the file."""
    items = read_file(file_name)
    items[str(user_id)] = username
    write_file(file_name, items)

def remove_from_file(file_name, user_id):
    """Remove an entry from the file."""
    items = read_file(file_name)
    if str(user_id) in items:
        del items[str(user_id)]
        write_file(file_name, items)

def is_admin(user_id):
    """Check if a user is an admin."""
    admins = read_file("admins.txt")
    return str(user_id) in admins

def is_admin_plus(user_id):
    """Check if a user is a bot admin."""
    admin_plus = read_file("admin_plus.txt")
    return str(user_id) in admin_plus

def get_allowed_games():
    """Read allowed game names from a text file."""
    if not os.path.exists("allowed_games.txt"):
        return []
    with open("allowed_games.txt", "r") as file:
        return [line.strip() for line in file if line.strip()]

async def generate_report_file(guild):
    """
    Generate an Excel report of all logged hours for all games and all users in the guild.
    Returns the file path of the saved report.
    """
    members = guild.members
    file_path = "weekly_hours_report.xlsx"

    async with aiosqlite.connect('practice_logger.db') as db:
        # Get all distinct games
        query_games = "SELECT DISTINCT game FROM hours"
        async with db.execute(query_games) as cursor:
            games = [row[0] for row in await cursor.fetchall()]

        # Initialize user data structure
        user_data = {m.id: {"name": m.display_name, "games": {}, "total_hours": 0} for m in members}

        # Fetch summed hours
        query_hours = '''
            SELECT user_id, game, SUM(hours) as total_hours
            FROM hours
            GROUP BY user_id, game
        '''
        async with db.execute(query_hours) as cursor:
            logs = await cursor.fetchall()

        # Populate user_data
        for user_id, game, total_hours in logs:
            if user_id in user_data:
                user_data[user_id]["games"][game] = total_hours
                user_data[user_id]["total_hours"] += total_hours

        # Create Workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # remove default sheet

        # Create summary sheet
        summary_sheet = workbook.create_sheet(title="Summary")
        summary_sheet.append(["User", "UserID", "Games Played", "Game Hours"])

        for uid, data in user_data.items():
            games_played = ", ".join(data["games"].keys()) if data["games"] else "No games"
            hours_str = ", ".join([f"{hours:.2f}" for hours in data["games"].values()]) if data["games"] else "0.00"
            summary_sheet.append([data["name"], uid, games_played, hours_str])

        # Conditional formatting in summary sheet
        # We apply formatting based on average of the user's logged hours (if multiple)
        for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=4, max_col=4):
            totals_cell = row[0]
            totals = totals_cell.value
            if totals and totals != "0.00":
                values = [float(val) for val in totals.split(", ")]
                avg = sum(values)/len(values)
                if avg >= 10:
                    fill = PatternFill(start_color="90EE90", fill_type="solid")  # Green
                elif 5 <= avg < 10:
                    fill = PatternFill(start_color="FFFFE0", fill_type="solid")  # Yellow
                else:
                    fill = PatternFill(start_color="FFC0CB", fill_type="solid")  # Red
                totals_cell.fill = fill

        # Create individual sheets per game
        for game in games:
            sheet = workbook.create_sheet(title=game)
            sheet.append(["User", "UserID", "Total Hours"])
            for uid, data in user_data.items():
                if game in data["games"]:
                    val = data["games"][game]
                    sheet.append([data["name"], uid, val])

            # Conditional formatting in game sheets
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
                total_hours_cell = row[0]
                val = total_hours_cell.value
                if val is not None:
                    if val >= 10:
                        fill = PatternFill(start_color="90EE90", fill_type="solid")  # Green
                    elif 5 <= val < 10:
                        fill = PatternFill(start_color="FFFFE0", fill_type="solid")  # Yellow
                    else:
                        fill = PatternFill(start_color="FFC0CB", fill_type="solid")  # Red
                    total_hours_cell.fill = fill

        workbook.save(file_path)
        return file_path