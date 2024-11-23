# Import the start_server function
from server import start_server
import discord
from discord.ext import commands, tasks
from discord import app_commands
import aiosqlite
import asyncio
import os
from dotenv import load_dotenv
import csv
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import pytz
from datetime import datetime, timedelta

# Start the keep-alive server
# start_server()
#Don't need this at the moment because it's hosted locally

# Load environment variables
load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
GUILD_ID = 717936676886020166  # Replace with your server's ID as an integer
LOG_CHANNEL_ID = 803387116989055006  # Replace with your desired channel ID

# Set up logging
logging.basicConfig(level=logging.INFO)

# Define intents for the bot
intents = discord.Intents.default()
intents.message_content = True
intents.presences = True
intents.members = True

# Initialize bot without a command prefix since slash commands are used
bot = commands.Bot(command_prefix=None, intents=intents)

# Database setup function
async def init_db():
    """Initialize the database with necessary tables."""
    async with aiosqlite.connect('practice_logger.db') as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS admins (
                user_id INTEGER PRIMARY KEY
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS buttons (
                name TEXT PRIMARY KEY
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS hours (
                user_id INTEGER,
                date TEXT,
                hours REAL,
                details TEXT,
                PRIMARY KEY (user_id, date)
            )
        ''')
        await db.execute('''
            CREATE TABLE IF NOT EXISTS auto_track (
                user_id INTEGER PRIMARY KEY,
                enabled INTEGER
            )
        ''')
        await db.commit()

async def migrate_db_schema():
    """Migrate the database schema to include 'game' in the primary key."""
    async with aiosqlite.connect('practice_logger.db') as db:
        # Check if the schema migration is needed
        async with db.execute("PRAGMA table_info(hours)") as cursor:
            columns = await cursor.fetchall()
            if any(col[1] == 'game' for col in columns):
                # Schema is already updated
                return

        # Rename the old table
        await db.execute("ALTER TABLE hours RENAME TO hours_old")

        # Create the new table with 'game' as part of the primary key
        await db.execute('''
            CREATE TABLE hours (
                user_id INTEGER,
                date TEXT,
                game TEXT,
                hours REAL,
                details TEXT,
                PRIMARY KEY (user_id, date, game)
            )
        ''')

        # Migrate data from the old table
        async with db.execute("SELECT user_id, date, details, hours FROM hours_old") as cursor:
            rows = await cursor.fetchall()
            for row in rows:
                user_id, date, details, hours = row
                game, additional_info = details.split(": ", 1) if ": " in details else (details, "")
                await db.execute('''
                    INSERT INTO hours (user_id, date, game, hours, details)
                    VALUES (?, ?, ?, ?, ?)
                ''', (user_id, date, game.strip(), hours, additional_info.strip()))

        # Drop the old table
        await db.execute("DROP TABLE hours_old")

        # Commit changes
        await db.commit()

# Global variables to store channel ID and message ID
daily_info = {"channel_id": None, "message_id": None}

def is_admin(user):
    """Check if a user is a admin."""
    admins = read_file("admins.txt")
    return str(user) in admins

def is_admin_plus(user):
    """Check if a user is a bot admin."""
    admin_plus = read_file("admin_plus.txt")
    return str(user) in admin_plus

def read_file(file_name):
    """Read admins from a text file."""
    if not os.path.exists(file_name):
        return {}
    with open(file_name, "r") as file:
        return {
            line.split()[0]: " ".join(line.split()[1:])
            for line in file if line.strip()
        }


def write_file(file_name, admins):
    """Write admins to a text file."""
    with open(file_name, "w") as file:
        for user_id, username in admins.items():
            file.write(f"{user_id} {username}\n")


def add_to_file(file_name, user_id, username):
    """Add an admin to the file."""
    admins = read_file(file_name)
    admins[str(user_id)] = username
    write_file(file_name, admins)


def remove_from_file(file_name, user_id):
    """Remove an admin from the file."""
    admins = read_file(file_name)
    if str(user_id) in admins:
        del admins[str(user_id)]
        write_file(file_name, admins)

def read_file(file_name):
    """Read admins from a text file."""
    if not os.path.exists(file_name):
        return {}
    with open(file_name, "r") as file:
        return {
            line.split()[0]: " ".join(line.split()[1:])
            for line in file if line.strip()
        }

async def schedule_daily_buttons():
    """Schedules the dailyButtons task to run at midnight CST."""
    while True:
        cst = pytz.timezone("America/Chicago")
        now = datetime.now(cst)
        next_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        sleep_duration = (next_midnight - now).total_seconds()
        await asyncio.sleep(sleep_duration) #sleep_duration
        await dailyButtons()

@bot.event
async def on_ready():
    """Triggered when the bot is ready."""
    await migrate_db_schema()  # Migrate the database schema if needed
    await init_db()  # Initialize the database
    try:
        bot.tree.copy_global_to(guild=discord.Object(id=GUILD_ID))
        synced = await bot.tree.sync(guild=discord.Object(id=GUILD_ID))
        print(f'Synced {len(synced)} commands to guild {GUILD_ID}')
    except Exception as e:
        print(f'Error syncing commands: {e}')
    
    # Register persistent views
    bot.add_view(DayButtonView([]))  # Empty initialization for registration

    bot.loop.create_task(schedule_daily_buttons()) #Begin the daily button loop

    print(f'{bot.user} has connected to Discord!')
    log_channel = bot.get_channel(LOG_CHANNEL_ID)
    if log_channel:
        await log_channel.send("I'm awake!")

@bot.tree.command(name="adddaily", description="Add a dailyButtons logic manually with a specific day.")
@app_commands.describe(day="The day to simulate (e.g., Monday, Tuesday).")
async def test_daily(interaction: discord.Interaction, day: str):
    """Manually trigger the dailyButtons logic for a specific day."""
    valid_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    if day not in valid_days:
        await interaction.response.send_message(f"Invalid day. Choose from: {', '.join(valid_days)}", ephemeral=True)
        return

    await dailyButtons(test_day=day)
    await interaction.response.send_message(f"Daily buttons logic executed for {day}.", ephemeral=True)


# Command: Ping
@bot.tree.command(name='ping', description='Check the bot\'s latency.')
async def ping(interaction: discord.Interaction):
    latency = round(bot.latency * 1000)
    await interaction.response.send_message(f'Pong! Latency: {latency}ms')

@bot.tree.command(name="sendembed", description="Send an embedded message to a specified channel.")
@app_commands.describe(
    channel_id="The channel ID to send the message to (optional). Defaults to the current channel.",
    title="The title of the embed.",
    description="The description of the embed."
)
async def make_embed(interaction: discord.Interaction, title: str, description: str, channel_id: str = None):
    """Send an embedded message to a specified channel."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return
    
    # Resolve the channel
    try:
        target_channel = bot.get_channel(int(channel_id)) if channel_id else interaction.channel
        if not target_channel:
            await interaction.response.send_message("Invalid or inaccessible channel.", ephemeral=True)
            return

        # Create and send the embed
        embed = discord.Embed(title=title, description=description, color=discord.Color.blue())
        await target_channel.send(embed=embed)

        # Respond to the command once
        await interaction.response.send_message(f"Embed sent to {target_channel.mention}.", ephemeral=True)
    except Exception as e:
        # Handle unexpected errors
        await interaction.response.send_message(f"An error occurred: {str(e)}", ephemeral=True)

@bot.tree.command(name="commands", description="List all available commands categorized by roles.")
@app_commands.describe(visible="Make the output visible to everyone? Defaults to False.")
async def list_commands(interaction: discord.Interaction, visible: bool=False):
    """
    List all available commands categorized by Admin+, Admin, and Normal.
    Reads from a 'commands' text file.
    """
    if not os.path.exists("commands.txt"):
        await interaction.response.send_message(
            "The commands file is missing. Please create a 'commands' file with the necessary data.",
            ephemeral=True,
        )
        return

    try:
        # Read the commands file
        with open("commands.txt", "r") as file:
            lines = [line.strip() for line in file if line.strip()]

        # Parse the commands into categories
        admin_plus_commands = []
        admin_commands = []
        normal_commands = []

        current_category = None
        for line in lines:
            if line.lower() == "[admin+]":
                current_category = admin_plus_commands
            elif line.lower() == "[admin]":
                current_category = admin_commands
            elif line.lower() == "[normal]":
                current_category = normal_commands
            else:
                if current_category is not None:
                    current_category.append(line)

        # Create the embed
        embed = discord.Embed(
            title="Available Commands",
            description="Here is a list of all available commands categorized by roles. Most commands have the option of making it visible.",
            color=discord.Color.blue(),
        )

        if admin_plus_commands:
            embed.add_field(
                name="**Admin+ Commands**",
                value="\n".join(admin_plus_commands),
                inline=False,
            )
        else:
            embed.add_field(
                name="**Admin+ Commands**",
                value="No commands found.",
                inline=False,
            )

        if admin_commands:
            embed.add_field(
                name="**Admin Commands**",
                value="\n".join(admin_commands),
                inline=False,
            )
        else:
            embed.add_field(
                name="**Admin Commands**",
                value="No commands found.",
                inline=False,
            )

        if normal_commands:
            embed.add_field(
                name="**Normal Commands**",
                value="\n".join(normal_commands),
                inline=False,
            )
        else:
            embed.add_field(
                name="**Normal Commands**",
                value="No commands found.",
                inline=False,
            )

        # Send the embed
        await interaction.response.send_message(embed=embed, ephemeral=not visible)

    except Exception as e:
        print(f"Error in /commands: {e}")
        await interaction.response.send_message(
            "An error occurred while reading the commands file. Please try again later.",
            ephemeral=True,
        )


@bot.tree.command(name="adminlist", description="List all bot and admins.")
@app_commands.describe(visible="Make the output visible to everyone? Defaults to False.")
async def adminlist(interaction: discord.Interaction, visible: bool = False):
    """List all Admin+ and admins."""

    admin_plus = read_file("admin_plus.txt")
    admins = read_file("admins.txt")

    embed = discord.Embed(
        title="Admin List",
        description="List of Admin+ and admins.",
        color=discord.Color.blue(),
    )

    if admin_plus:
        admin_plus_list = "\n".join([f"{username} (ID: {user_id})" for user_id, username in admin_plus.items()])
        embed.add_field(name="Admin+", value=admin_plus_list, inline=False)
    else:
        embed.add_field(name="Admin+", value="No Admin+ found.", inline=False)

    if admins:
        admins_list = "\n".join([f"{username} (ID: {user_id})" for user_id, username in admins.items()])
        embed.add_field(name="Admins", value=admins_list, inline=False)
    else:
        embed.add_field(name="Admins", value="No admins found.", inline=False)

    # Send the embed, respecting the visibility option
    await interaction.response.send_message(embed=embed, ephemeral=not visible)


@bot.tree.command(name="removeadmin", description="Remove a admin from the bot (Admin+ only).")
@app_commands.describe(user="The user to be removed as a admin.")
async def removeadmin(interaction: discord.Interaction, user: discord.User):
    """Remove a admin from the admin list."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    #Check is user is trying to remove themselves
    if user.id == interaction.user.id:
        await interaction.response.send_message("You can't remove yourself from the admin list!", ephemeral=True)
        return

    remove_from_file("admins.txt", user.id)

    # Log action
    log_channel = bot.get_channel(LOG_CHANNEL_ID)
    if log_channel:
        await log_channel.send(f"{interaction.user.display_name} removed {user.display_name} from admin roles.")

    await interaction.response.send_message(f"{user.display_name} has been removed as a admin.", ephemeral=True)


#Command: Admin Add - adds an admin to the file
@bot.tree.command(name="addadmin", description="Add a admin to the bot (Admin+ only).")
@app_commands.describe(user="The user to be added as a admin.")
async def addadmin(interaction: discord.Interaction, user: discord.User):
    """Add a admin to the admin list."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    add_to_file("admins.txt", user.id, user.name)

    # Log action
    log_channel = bot.get_channel(LOG_CHANNEL_ID)
    if log_channel:
        await log_channel.send(f"{interaction.user.display_name} added {user.display_name} as a admin.")

    await interaction.response.send_message(f"{user.display_name} has been added as a admin.", ephemeral=True)

#Command:Bot Admin Remove - removes a bot admin from the file
@bot.tree.command(name="removeadminplus", description="Remove a bot admin from the bot (Admin+ only).")
@app_commands.describe(user="The user to be removed as a bot admin.")
async def removeadmin(interaction: discord.Interaction, user: discord.User):
    """Remove a admin from the admin list."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    #Check is user is trying to remove themselves
    if user.id == interaction.user.id:
        await interaction.response.send_message("You can't remove yourself from the admin list!", ephemeral=True)
        return

    remove_from_file("admin_plus.txt", user.id)

    # Log action
    log_channel = bot.get_channel(LOG_CHANNEL_ID)
    if log_channel:
        await log_channel.send(f"{interaction.user.display_name} removed {user.display_name} from bot admin roles.")

    await interaction.response.send_message(f"{user.display_name} has been removed as a bot admin.", ephemeral=True)


#Command: Bot Admin Add - adds a bot admin to the file
@bot.tree.command(name="addadminplus", description="Add a bot admin to the bot (Admin+ only).")
@app_commands.describe(user="The user to be added as a bot admin.")
async def adminadd(interaction: discord.Interaction, user: discord.User):
    """Add a admin to the admin list."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    add_to_file("admin_plus.txt", user.id, user.name)

    # Log action
    log_channel = bot.get_channel(LOG_CHANNEL_ID)
    if log_channel:
        await log_channel.send(f"{interaction.user.display_name} added {user.display_name} as a bot admin.")

    await interaction.response.send_message(f"{user.display_name} has been added as a bot admin.", ephemeral=True)


# Command: Reset Week
@bot.tree.command(name="resetweek", description="Reset all logged hours for all users (Admin+ only).")
async def resetweek(interaction: discord.Interaction):
    """Reset all logged hours for the current week."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    try:
        async with aiosqlite.connect('practice_logger.db') as db:
            # Delete all entries in the 'hours' table
            await db.execute("DELETE FROM hours")
            await db.commit()

        # Log action
        log_channel = bot.get_channel(LOG_CHANNEL_ID)
        if log_channel:
            await log_channel.send(f"{interaction.user.display_name} reset all logged hours for the week.")

        await interaction.response.send_message("All logged hours for the week have been reset.", ephemeral=True)
    except Exception as e:
        print(f"Error in /resetweek command: {e}")
        await interaction.response.send_message("An error occurred while resetting the week. Please try again later.", ephemeral=True)

# Utility Function to Read Allowed Games from File
def get_allowed_games():
    """Read allowed game names from a text file."""
    if not os.path.exists("allowed_games.txt"):
        return []  # Return an empty list if the file doesn't exist
    with open("allowed_games.txt", "r") as file:
        games = [line.strip() for line in file if line.strip()]  # Strip empty lines and whitespace
    return games

async def add_hours_to_db(user_id: int, game: str, date: str, hours: float, additional_info: str = None):
    """Add hours to the database, consolidating entries for the same user, game, and day."""
    async with aiosqlite.connect('practice_logger.db') as db:
        # Check if an entry already exists
        query_check = '''
            SELECT hours
            FROM hours
            WHERE user_id = ? AND date = ? AND game = ?
        '''
        async with db.execute(query_check, (user_id, date, game)) as cursor:
            existing_entry = await cursor.fetchone()

        if existing_entry:
            # Update the entry by adding the new hours
            current_hours = existing_entry[0]
            new_hours = current_hours + hours
            query_update = '''
                UPDATE hours
                SET hours = ?, details = ?
                WHERE user_id = ? AND date = ? AND game = ?
            '''
            await db.execute(query_update, (new_hours, additional_info, user_id, date, game))
        else:
            # Insert a new entry
            query_insert = '''
                INSERT INTO hours (user_id, date, game, hours, details)
                VALUES (?, ?, ?, ?, ?)
            '''
            await db.execute(query_insert, (user_id, date, game, hours, additional_info or "No additional information"))

        await db.commit()


class LogHoursModal(discord.ui.Modal):
    """A modal to log game hours."""

    def __init__(self, day: str):
        super().__init__(title=f"Log Hours for {day}")
        self.day = day

        # Game Name (Required)
        self.game_name = discord.ui.TextInput(
            label="Game Name",
            placeholder="Enter the game's name (e.g., League of Legends)",
            required=True,
            max_length=100,
        )
        self.add_item(self.game_name)

        # Hours Played Today (Required)
        self.hours_played = discord.ui.TextInput(
            label="Hours Played Today",
            placeholder="Enter the number of hours (e.g., 2.5)",
            required=True,
            max_length=10,
        )
        self.add_item(self.hours_played)

        # Additional Information (Optional)
        self.additional_info = discord.ui.TextInput(
            label="Additional Information",
            placeholder="Any extra details (optional)",
            required=False,
            style=discord.TextStyle.paragraph,
        )
        self.add_item(self.additional_info)

    async def on_submit(self, interaction: discord.Interaction):
        """Handle the modal submission and save data to the database."""
        game_name = self.game_name.value.strip()
        hours_input = self.hours_played.value.strip()
        additional_info = self.additional_info.value.strip() or "No additional information provided."

        # Validate game name
        allowed_games = get_allowed_games()
        if game_name not in allowed_games:
            await interaction.response.send_message(
                f"Invalid game name. Allowed games are: {', '.join(allowed_games)}.",
                ephemeral=True,
            )
            return

        # Validate hours as float
        try:
            hours = float(hours_input)
            if hours <= 0:
                raise ValueError("Hours must be positive.")
        except ValueError:
            await interaction.response.send_message(
                "Invalid input for hours. Enter a positive number (e.g., 2.5).",
                ephemeral=True,
            )
            return

        try:
            today = datetime.now(pytz.timezone("America/Chicago"))
            start_of_week = today - timedelta(days=today.weekday())
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            day_index = days.index(self.day)
            target_date = start_of_week + timedelta(days=day_index)
            target_date_str = target_date.strftime('%Y-%m-%d')

            # Use the helper function to add hours
            await add_hours_to_db(interaction.user.id, game_name, target_date_str, hours, additional_info)

            # Log to the channel
            log_channel = bot.get_channel(LOG_CHANNEL_ID)
            if log_channel:
                await log_channel.send(
                    f"{interaction.user.display_name} has added {hours} hours for {game_name} on {self.day}."
                )

            # Notify the user
            await interaction.response.send_message(
                f"**Logged Hours for {self.day}:**\n"
                f"- **Game Name:** {game_name}\n"
                f"- **Hours Played Today:** {hours}\n"
                f"- **Additional Information:** {additional_info}\n\n"
                f"The data has been saved successfully!",
                ephemeral=True
            )
        except Exception as e:
            print(f"Error saving data to database: {e}")
            await interaction.response.send_message(
                "An error occurred while saving your data. Please try again later.",
                ephemeral=True,
            )


            # Get the target date
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            today = datetime.now(pytz.timezone("America/Chicago"))
            start_of_week = today - timedelta(days=today.weekday())
            day_index = days.index(self.day)
            target_date = start_of_week + timedelta(days=day_index)
            target_date_str = target_date.strftime('%Y-%m-%d')

            # Consolidate hours
            await add_hours_to_db(interaction.user.id, game_name, target_date_str, hours, additional_info)

            await interaction.response.send_message(
                f"Logged {hours} hours for {game_name} on {self.day}.\nAdditional Info: {additional_info}",
                ephemeral=True
            )
        except ValueError:
            await interaction.response.send_message("Invalid input for hours. Please enter a positive number.", ephemeral=True)
        except Exception as e:
            print(f"Error in LogHoursModal: {e}")
            await interaction.response.send_message("An error occurred. Please try again.", ephemeral=True)
    
#Setup the daily channelid and messageid
@bot.tree.command(name="setdailyinfo", description="Set up the channel ID and message ID for daily button management.")
@app_commands.describe(channel_id="The ID of the channel where the message is located.", message_id="The ID of the message to manage.")
async def set_daily_info(interaction: discord.Interaction, channel_id: str, message_id: str):
    """Command to set the channel ID and message ID for dailyButtons."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    try:
        # Update the global daily_info dictionary
        daily_info["channel_id"] = int(channel_id)
        daily_info["message_id"] = int(message_id)

        await interaction.response.send_message(
            f"Daily button setup updated! Channel ID: `{channel_id}`, Message ID: `{message_id}`", ephemeral=True
        )
    except ValueError:
        await interaction.response.send_message("Invalid ID format. Please provide numeric IDs.", ephemeral=True)

# Daily action logic
async def dailyButtons(test_day=None):
    if not daily_info["channel_id"] or not daily_info["message_id"]:
        print("Daily info not set. Use /setdailyinfo to configure channel and message IDs.")
        return

    cst = pytz.timezone("America/Chicago")
    today = test_day or datetime.now(cst).strftime("%A")  # Get the current day (e.g., Monday, Tuesday)

    channel = bot.get_channel(daily_info["channel_id"])
    if not channel:
        print("Channel not found.")
        return

    match today:
        case "Monday":
            await clearbuttons(daily_info["message_id"], channel)
            await addbuttons(daily_info["message_id"], channel, "Monday")
        case "Tuesday":
            await addbuttons(daily_info["message_id"], channel, "Tuesday")
        case "Wednesday":
            await addbuttons(daily_info["message_id"], channel, "Wednesday")
        case "Thursday":
            await addbuttons(daily_info["message_id"], channel, "Thursday")
        case "Friday":
            await addbuttons(daily_info["message_id"], channel, "Friday")
        case "Saturday":
            await addbuttons(daily_info["message_id"], channel, "Saturday")
        case "Sunday":
            await addbuttons(daily_info["message_id"], channel, "Sunday")


@bot.tree.command(name="games", description="Show the list of allowed games.")
@app_commands.describe(visible="Make the output visible to everyone? Defaults to False.")
async def games(interaction: discord.Interaction, visible: bool = False):
    """Command to display the allowed games list."""
    allowed_games = get_allowed_games()

    if not allowed_games:
        await interaction.response.send_message(
            "No allowed games have been defined. Please contact an admin.",
            ephemeral=not visible,  # Respect the visibility option
        )
        return

    # Create an embed for the allowed games list
    embed = discord.Embed(
        title="Allowed Games",
        description="\n".join([f"- {game}" for game in allowed_games]),
        color=discord.Color.blue()
    )
    embed.set_footer(text="Use these exact names to log your hours.")

    # Send the embed, respecting the visibility option
    await interaction.response.send_message(embed=embed, ephemeral=not visible)


@bot.tree.command(name="clearbuttons", description="Remove all buttons from a message (admin only).")
@app_commands.describe(message_id="The ID of the message to clear buttons from.")
async def clearbuttons(interaction: discord.Interaction, message_id: str):
    """Remove all buttons from a message."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message("You are not authorized to use this command.", ephemeral=True)
        return

    try:
        # Fetch the message
        message = await interaction.channel.fetch_message(int(message_id))
        await message.edit(view=None)


        await interaction.response.send_message(
            f"Successfully removed all buttons from the message with ID `{message_id}`.", ephemeral=True
        )

    except discord.NotFound:
        await interaction.response.send_message("Message not found. Please check the message ID.", ephemeral=True)
    except discord.Forbidden:
        await interaction.response.send_message(
            "I do not have permission to edit this message. Ensure I have the necessary permissions.", ephemeral=True
        )
    except discord.HTTPException as e:
        await interaction.response.send_message(
            f"An error occurred while trying to clear buttons: {str(e)}", ephemeral=True
        )


@bot.tree.command(name="reset", description="Reset hours for a user for a specific day or the full week (admin only).")
@app_commands.describe(
    day="The day to reset (leave blank to reset the full week).",
    user="The user whose hours to reset."
)
async def reset(interaction: discord.Interaction, user: discord.User, day: str = None):
    """Reset hours for a user for a specific day or the full week."""
    if not is_admin(interaction.user.id):
        await interaction.response.send_message("You are not authorized to reset hours.", ephemeral=True)
        return

    admin_name = interaction.user.display_name

    try:
        async with aiosqlite.connect('practice_logger.db') as db:
            today = datetime.now(pytz.timezone("America/Chicago"))
            start_of_week = today - timedelta(days=today.weekday())

            if day:
                days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                if day.capitalize() not in days:
                    await interaction.response.send_message("Invalid day. Please specify a valid day of the week.", ephemeral=True)
                    return

                day_index = days.index(day.capitalize())
                target_date = start_of_week + timedelta(days=day_index)
                target_date_str = target_date.strftime('%Y-%m-%d')

                # Delete all games logged on the specific day
                await db.execute("DELETE FROM hours WHERE user_id = ? AND date = ?", (user.id, target_date_str))
                await db.commit()

                await interaction.response.send_message(f"Reset hours for {user.display_name} on {day}.", ephemeral=True)
                log_message = f"{admin_name} reset hours for {user.display_name} on {day}."
            else:
                start_of_week_str = start_of_week.strftime('%Y-%m-%d')
                await db.execute("DELETE FROM hours WHERE user_id = ? AND date >= ?", (user.id, start_of_week_str))
                await db.commit()

                await interaction.response.send_message(f"Reset hours for {user.display_name} for the full week.", ephemeral=True)
                log_message = f"{admin_name} reset hours for {user.display_name} for the full week."

        log_channel = bot.get_channel(LOG_CHANNEL_ID)
        if log_channel:
            await log_channel.send(log_message)
    except Exception as e:
        print(f"Error in /reset command: {e}")
        await interaction.response.send_message("An error occurred while resetting hours. Please try again later.", ephemeral=True)


@bot.tree.command(name="hours", description="View your logged hours for the week.")
@app_commands.describe(
    user="Optional: Specify a user to view their logged hours (Admin Only).",
    visible="Make the output visible to everyone? Defaults to False."
)
async def hours(interaction: discord.Interaction, user: discord.User = None, visible: bool = False):
    """Command to show logged hours for the week."""
    target_user = user or interaction.user

    # Defer the response with visibility option
    await interaction.response.defer(ephemeral=not visible)

    if not user == None and not user == interaction.user and not is_admin(interaction.user.id):
        await interaction.followup.send("You aren't able to view someone else's hours!", ephemeral=True)
        return

    try:
        async with aiosqlite.connect('practice_logger.db') as db:
            # Start of the week (Monday)
            today = datetime.now(pytz.timezone("America/Chicago"))
            start_of_week = today - timedelta(days=today.weekday())
            start_of_week_str = start_of_week.strftime('%Y-%m-%d')

            # Fetch logs for the week
            query = '''
                SELECT date, game, SUM(hours) as total_hours, details
                FROM hours
                WHERE user_id = ? AND date >= ?
                GROUP BY date, game, details
                ORDER BY date, game
            '''
            async with db.execute(query, (target_user.id, start_of_week_str)) as cursor:
                logs = await cursor.fetchall()

            # Organize logs by day of the week
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            logs_by_day = {day: [] for day in days}
            weekly_totals = {}

            for log_date, game, hours, details in logs:
                log_date_obj = datetime.strptime(log_date, '%Y-%m-%d')
                day_name = log_date_obj.strftime('%A')

                # Add detailed log for the day
                logs_by_day[day_name].append(f"{game}: {hours:.2f} hours - {details}")

                # Consolidate weekly totals by game
                weekly_totals[game] = weekly_totals.get(game, 0) + hours

            # Create the embed
            embed = discord.Embed(
                title=f"Weekly Logged Hours for {target_user.display_name}",
                description="Here are the hours logged this week, separated by day and game.",
                color=discord.Color.blue(),
                timestamp=datetime.now()
            )

            # Add logs to embed, day by day
            for day, logs in logs_by_day.items():
                if logs:
                    embed.add_field(name=f"**{day}**", value="\n".join(logs), inline=False)
                else:
                    embed.add_field(name=f"**{day}**", value="No hours logged", inline=False)

            # Weekly totals section
            weekly_totals_formatted = "\n".join([f"{game}: **{total:.2f} hours**" for game, total in weekly_totals.items()])
            embed.add_field(
                name="**Weekly Totals by Game**",
                value=weekly_totals_formatted or "No hours logged this week.",
                inline=False
            )

        # Send the response with visibility based on the `visible` option
        await interaction.followup.send(embed=embed, ephemeral=not visible)
    except Exception as e:
        print(f"Error in /hours command: {e}")
        await interaction.followup.send(
            "An error occurred while fetching your logged hours. Please try again later.",
            ephemeral=True,
        )



@bot.tree.command(name="add", description="Add hours to a specific game for a user on a specific day (admin only).")
@app_commands.describe(
    day="The day to log the hours (e.g., Monday).",
    game="The game to add hours to.",
    hours="The number of hours to add.",
    user="The user to add hours for."
)
async def add(
    interaction: discord.Interaction,
    day: str,
    game: str,
    hours: float,
    user: discord.User
):
    """Add hours to a specific game for a user on a specific day."""
    if not is_admin(interaction.user.id):
        await interaction.response.send_message("You are not authorized to add hours.", ephemeral=True)
        return

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    if day.capitalize() not in days:
        await interaction.response.send_message("Invalid day. Please specify a valid day of the week.", ephemeral=True)
        return

    admin_name = interaction.user.display_name

    try:
        today = datetime.now(pytz.timezone("America/Chicago"))
        start_of_week = today - timedelta(days=today.weekday())
        day_index = days.index(day.capitalize())
        target_date = start_of_week + timedelta(days=day_index)
        target_date_str = target_date.strftime('%Y-%m-%d')

        # Use the helper function to add hours
        await add_hours_to_db(user.id, game, target_date_str, hours, f"Added by {admin_name}")

        await interaction.response.send_message(f"Added {hours} hours for {user.display_name} on {game} for {day}.", ephemeral=True)

        # Log action to the log channel
        log_channel = bot.get_channel(LOG_CHANNEL_ID)
        if log_channel:
            await log_channel.send(f"{admin_name} added {hours} hours to {user.display_name} for {game} on {day}.")
    except Exception as e:
        print(f"Error in /add command: {e}")
        await interaction.response.send_message("An error occurred while adding hours. Please try again later.", ephemeral=True)


@bot.tree.command(name="remove", description="Remove hours from a specific game for a user on a specific day (admin only).")
@app_commands.describe(
    day="The day to remove the hours from (e.g., Monday).",
    game="The game to remove hours from.",
    hours="The number of hours to remove.",
    user="The user to remove hours for."
)
async def remove(interaction: discord.Interaction, day: str, game: str, hours: float, user: discord.User):
    """Remove hours from a specific game for a user on a specific day."""
    if not is_admin(interaction.user.id):
        await interaction.response.send_message("You are not authorized to remove hours.", ephemeral=True)
        return

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    if day.capitalize() not in days:
        await interaction.response.send_message("Invalid day. Please specify a valid day of the week.", ephemeral=True)
        return

    admin_name = interaction.user.display_name

    try:
        today = datetime.now(pytz.timezone("America/Chicago"))
        start_of_week = today - timedelta(days=today.weekday())
        day_index = days.index(day.capitalize())
        target_date = start_of_week + timedelta(days=day_index)
        target_date_str = target_date.strftime('%Y-%m-%d')

        async with aiosqlite.connect('practice_logger.db') as db:
            # Check for existing hours for the user, date, and game
            query = '''
                SELECT hours
                FROM hours
                WHERE user_id = ? AND date = ? AND game = ?
            '''
            async with db.execute(query, (user.id, target_date_str, game)) as cursor:
                row = await cursor.fetchone()

            if not row:
                await interaction.response.send_message(
                    f"No hours found for {game} on {day} for {user.display_name}.",
                    ephemeral=True
                )
                return

            existing_hours = row[0]
            if existing_hours < hours:
                await interaction.response.send_message(
                    f"{user.display_name} does not have enough hours logged on {day} for {game} to remove.",
                    ephemeral=True
                )
                return

            # Calculate new hours
            new_hours = existing_hours - hours
            if new_hours == 0:
                # If no hours are left, delete the entry
                query_delete = '''
                    DELETE FROM hours
                    WHERE user_id = ? AND date = ? AND game = ?
                '''
                await db.execute(query_delete, (user.id, target_date_str, game))
            else:
                # Update the entry with the new hours
                query_update = '''
                    UPDATE hours
                    SET hours = ?
                    WHERE user_id = ? AND date = ? AND game = ?
                '''
                await db.execute(query_update, (new_hours, user.id, target_date_str, game))

            # Commit the changes
            await db.commit()

        # Notify the admin and log the action
        await interaction.response.send_message(
            f"Removed {hours} hours for {user.display_name} on {game} for {day}.",
            ephemeral=True
        )

        log_channel = bot.get_channel(LOG_CHANNEL_ID)
        if log_channel:
            await log_channel.send(
                f"{admin_name} removed {hours} hours from {user.display_name} for {game} on {day}."
            )
    except Exception as e:
        print(f"Error in /remove command: {e}")
        await interaction.response.send_message("An error occurred while removing hours. Please try again later.", ephemeral=True)


# Add Weekly Buttons
class DayButtonView(discord.ui.View):
    """A view containing buttons for each day of the week."""
    def __init__(self, days):
        super().__init__(timeout=None)
        for day in days:
            button = discord.ui.Button(label=day, style=discord.ButtonStyle.primary, custom_id=f"day_{day.lower()}")
            button.callback = self.button_callback  # Attach the callback dynamically
            self.add_item(button)

    async def button_callback(self, interaction: discord.Interaction):
        """Callback for button interactions."""
        # Extract custom_id from interaction data
        custom_id = interaction.data.get("custom_id", "")
        if custom_id.startswith("day_"):
            day = custom_id.split('_')[1].capitalize()
            
            # Trigger the modal
            modal = LogHoursModal(day)
            await interaction.response.send_modal(modal)

# Add buttons for a specific day
async def addbuttons(message_id, channel, new_day):
    """Add a button for the new day without removing existing buttons."""
    try:
        # Fetch the message
        message = await channel.fetch_message(int(message_id))

        # Get current buttons if any
        if message.components:  # Check if there are already buttons
            existing_days = [
                component.label
                for row in message.components
                for component in row.children
            ]
        else:
            existing_days = []

        # Add the new day to the list
        if new_day not in existing_days:
            existing_days.append(new_day)

        # Create a new view with the updated days
        view = DayButtonView(existing_days)
        await message.edit(view=view)

    except discord.NotFound:
        print(f"Message with ID {message_id} not found.")
    except discord.Forbidden:
        print("Bot lacks permissions to edit the message.")
    except discord.HTTPException as e:
        print(f"An HTTP exception occurred: {e}")

# Clear buttons from the specified message
async def clearbuttons(message_id, channel):
    try:
        message = await channel.fetch_message(int(message_id))
        await message.edit(view=None)
    except discord.NotFound:
        print(f"Message with ID {message_id} not found.")
    except discord.Forbidden:
        print("Bot lacks permissions to edit the message.")
    except discord.HTTPException as e:
        print(f"An HTTP exception occurred: {e}")

@bot.tree.command(name='addbuttons', description='Add buttons to a message for days of the week (starting from Monday).')
@app_commands.describe(message_id='The ID of the message to add buttons to.')
async def add_buttons(interaction: discord.Interaction, message_id: str):
    """Add buttons dynamically for the days of the week, starting from Monday."""
    if not is_admin_plus(interaction.user.id):
        await interaction.response.send_message('You are not authorized to use this command.', ephemeral=True)
        return

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']  # Default to Monday
    view = DayButtonView(days)
    channel = interaction.channel

    try:
        message = await interaction.channel.fetch_message(message_id)
        await message.edit(view=view)
        await interaction.response.send_message(f'Buttons for days starting from Monday have been added to message {message_id}.', ephemeral=True)
    except discord.NotFound:
        await interaction.response.send_message('Message not found. Please check the message ID.', ephemeral=True)
    except discord.Forbidden:
        await interaction.response.send_message('Bot lacks permissions to edit the message.', ephemeral=True)
    except discord.HTTPException as e:
        await interaction.response.send_message(f"An error occurred while adding buttons: {str(e)}", ephemeral=True)


@bot.tree.command(name="report", description="Generate an Excel report of logged hours for all games.")
@app_commands.describe(visible="Make the report visible to everyone? Defaults to False.")
async def report(interaction: discord.Interaction, visible: bool = False):
    """Generate a report of all logged hours into an Excel file."""
    await interaction.response.defer(ephemeral=not visible)  # Acknowledge the command

    if not is_admin(interaction.user.id):
        await interaction.followup.send("You are not authorized to generate reports.", ephemeral=True)
        return

    try:
        # Fetch all members and roles in the guild
        guild = interaction.guild
        members = guild.members

        async with aiosqlite.connect('practice_logger.db') as db:
            # Fetch distinct games from the database
            query_games = "SELECT DISTINCT game FROM hours"
            async with db.execute(query_games) as cursor:
                games = [row[0] for row in await cursor.fetchall()]

            # Prepare user data dictionary
            user_data = {member.id: {"name": member.display_name, "games": {}, "total_hours": 0} for member in members}

            # Fetch logs for all users
            query_hours = '''
                SELECT user_id, game, SUM(hours) as total_hours
                FROM hours
                GROUP BY user_id, game
            '''
            async with db.execute(query_hours) as cursor:
                logs = await cursor.fetchall()

            # Populate the user data dictionary
            for log in logs:
                user_id = log[0]
                game_name = log[1]
                total_hours = log[2]

                if user_id in user_data:
                    user_data[user_id]["games"][game_name] = total_hours
                    user_data[user_id]["total_hours"] += total_hours

            # Prepare the Excel workbook
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet

            # Create a summary sheet for all users
            summary_sheet = workbook.create_sheet(title="Summary")
            summary_sheet.append(["User", "UserID", "Games Played", "Total Weekly Hours"])

            # Populate the summary sheet
            for user_id, data in user_data.items():
                games_played = list(data["games"].keys())
                game_totals = list(data["games"].values())

                row = [
                    data["name"],  # User
                    user_id,  # UserID
                    ", ".join(games_played) if games_played else "No games played",  # Games Played
                    ", ".join([f"{hours:.2f}" for hours in game_totals]) if game_totals else "0.00"  # Weekly Hours per Game
                ]
                summary_sheet.append(row)

            # Apply conditional formatting to the summary sheet
            for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, min_col=4, max_col=4):
                totals_cell = row[0]
                totals = totals_cell.value
                if totals and totals != "0.00":
                    for total in totals.split(", "):
                        total_hours = float(total)
                        if total_hours >= 10:
                            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                        elif 5 <= total_hours < 10:
                            fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Yellow
                        else:
                            fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Red
                        totals_cell.fill = fill

            # Create individual sheets per game
            for game in games:
                sheet = workbook.create_sheet(title=game)
                sheet.append(["User", "UserID", "Total Hours"])

                for user_id, data in user_data.items():
                    if game in data["games"]:
                        sheet.append([data["name"], user_id, data["games"][game]])

                # Apply conditional formatting to the game sheet
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
                    total_hours_cell = row[0]
                    total_hours = total_hours_cell.value
                    if total_hours is not None:
                        if total_hours >= 10:
                            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
                        elif 5 <= total_hours < 10:
                            fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Yellow
                        else:
                            fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Red
                        total_hours_cell.fill = fill
                        
            # Save the workbook
            file_path = "weekly_hours_report.xlsx"  # Save in the current directory
            workbook.save(file_path)

            # Send the file to the user or channel based on visibility
            await interaction.followup.send(
                content="Here is the compiled report:",
                file=discord.File(file_path),
                ephemeral=not visible  # Show to everyone if visible is True
            )

    except Exception as e:
        print(f"Error in /report command: {e}")
        await interaction.followup.send(
            content="An error occurred while generating the report. Please try again later.",
            ephemeral=True
        )

bot.run(TOKEN)
